"""Toplu Kayıt Aktarımı — kurumun Excel/CSV öğrenci-kur listesini sisteme aktarır.

Akış (sinav.py taslak-onay deseni): yukle (parse + normalize + TASLAK yaz) →
admin eşleştirme ekranında düzeltir (PUT) → uygula (idempotent kullanıcı/alacak
oluşturma; dry-run destekli). Hiçbir kullanıcı admin onayı olmadan oluşturulmaz.

Her SATIR bir öğrenci-KUR kaydıdır (öğrenci değil). Aynı öğrenci farklı kurlarla
birden çok satırda; aynı veli (telefon) birden çok öğrencide olabilir.

Koleksiyon: db.toplu_kayit_taslaklari  (durum: taslak | uygulandi | iptal)
Normalizasyon: core.kayit_normalize (saf fonksiyonlar).

Yollar (prefix=/api):
  POST   /toplu-kayit/yukle          (admin — xlsx/csv parse + taslak)
  GET    /toplu-kayit/taslak/{id}    (admin — eşleştirme ekranı verisi)
  PUT    /toplu-kayit/taslak/{id}    (admin — kolon eşleme/ücret/satır düzeltmeleri)
  POST   /toplu-kayit/uygula/{id}    (admin — dry_run destekli uygulama)
  GET    /toplu-kayit/rapor/{id}/{tur}.xlsx  (admin — şifre/hata raporu)
"""
import io
import csv
import uuid
import logging
from datetime import datetime

import re

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, Query
from fastapi.responses import StreamingResponse

from core.db import db
from core.auth import require_role, UserRole, hash_password
from core.hesap import gecici_sifre_uret, ogretmen_kaydi_olustur
from core import kayit_normalize as KN

router = APIRouter()

_YAZMA = require_role(UserRole.ADMIN, UserRole.COORDINATOR)
MAX_DOSYA_BYTE = 8 * 1024 * 1024  # 8 MB

# Varsayılan kolon eşlemesi (pozisyon bazlı — başlıklar anlamsız).
_VARSAYILAN_KOLON = {
    "kayit_tarihi": 0, "ogretmen_ad": 1, "ogrenci_ad": 2, "sinif": 3, "kur": 4,
    "veli_ad": 5, "veli_telefon": 7, "notlar": 8, "durum": 9,
}
# Başlık satırı tespiti için ipuçları.
_BASLIK_IPUCU = ["telefon", "öğretmen", "ogretmen", "öğrenci", "ogrenci", "sınıf", "sinif", "kur", "veli", "tarih", "not"]


def _hucre_str(c) -> str:
    """Hücre → temiz string. Excel tam-sayı float'ı ('5052627395.0'→'5052627395',
    '2.0'→'2') ve datetime ('2024-09-16 17:02:10') düzgün işlenir."""
    if c is None:
        return ""
    if isinstance(c, bool):
        return str(c)
    if isinstance(c, float) and c.is_integer():
        return str(int(c))
    if isinstance(c, datetime):
        return c.isoformat(sep=" ")
    return str(c).strip()


def _parse_dosya(icerik: bytes, dosya_adi: str, sayfa: str | None = None) -> tuple[list[list], list[str]]:
    """xlsx/csv → (satırlar, sayfa_adlari). xlsx'te `sayfa` verilirse o sayfa, yoksa ilk."""
    ad = (dosya_adi or "").lower()
    if ad.endswith(".xlsx") or ad.endswith(".xlsm"):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(icerik), read_only=True, data_only=True)
        sayfalar = list(wb.sheetnames)
        sn = sayfa if (sayfa and sayfa in sayfalar) else sayfalar[0]
        ws = wb[sn]
        satirlar = [[_hucre_str(c) for c in row] for row in ws.iter_rows(values_only=True)]
        wb.close()
        return satirlar, sayfalar
    if ad.endswith(".csv"):
        metin = icerik.decode("utf-8-sig", errors="replace")
        ornek = metin[:2000]
        ayirici = ";" if ornek.count(";") > ornek.count(",") else ","
        return [[str(c).strip() for c in row] for row in csv.reader(io.StringIO(metin), delimiter=ayirici)], []
    raise HTTPException(status_code=400, detail="Yalnızca .xlsx veya .csv dosyaları desteklenir")


def _baslik_mi(satir: list) -> bool:
    """İlk satır başlık mı (veri değil) — ipucu kelimelerinden en az 2'si geçiyorsa."""
    metin = " ".join(str(c) for c in satir).lower()
    return sum(1 for ip in _BASLIK_IPUCU if ip in metin) >= 2


def _hucre(satir: list, idx) -> str:
    if idx is None or idx < 0 or idx >= len(satir):
        return ""
    return str(satir[idx] or "").strip()


async def _ogretmen_listesi() -> list[dict]:
    docs = await db.teachers.find({"arsivli": {"$ne": True}}, {"_id": 0, "id": 1, "ad": 1, "soyad": 1}).to_list(length=2000)
    return docs


def _satir_isle(satir: list, kolon: dict, ogretmenler: list[dict], satir_no: int) -> dict:
    """Bir ham satırı normalize eder + kuyruk belirler. Çoklu kur = tek satır kaydı
    ama kur listesi (uygulama her kur için ayrı alacak açar)."""
    ham = {alan: _hucre(satir, idx) for alan, idx in kolon.items()}

    ogr = KN.normalize_ogrenci_ad(ham.get("ogrenci_ad", ""))
    tel = KN.normalize_telefon(ham.get("veli_telefon", ""))
    kurlar = KN.normalize_kur(ham.get("kur", ""))
    veli = KN.normalize_ad(ham.get("veli_ad", ""))
    veli_parca = veli.split(" ", 1)
    not_sinif = KN.siniflandir_not(ham.get("notlar", ""))
    ot = KN.ogretmen_eslestir(ham.get("ogretmen_ad", ""), ogretmenler) if ham.get("ogretmen_ad") else {"oneriler": [], "en_iyi": None, "otomatik": False}

    norm = {
        "kayit_tarihi": KN.normalize_tarih(ham.get("kayit_tarihi", "")),
        "ogrenci_ad": ogr["ad"], "ogrenci_soyad": ogr["soyad"],
        "ogrenci_gecerli": ogr["gecerli"], "ogrenci_sebep": ogr["sebep"],
        "sinif": KN.normalize_sinif(ham.get("sinif", "")),
        "kurlar": kurlar,
        "veli_ad": veli_parca[0] if veli_parca else "",
        "veli_soyad": veli_parca[1] if len(veli_parca) > 1 else "",
        "veli_telefon": tel["e164"], "veli_telefon_gecerli": tel["gecerli"],
        "odeme_durumu": not_sinif["odeme_durumu"],
        "egitim_notu": not_sinif["egitim_notu"],
        "aciklama": not_sinif["aciklama"],
        "taksit_notu": not_sinif["taksit_notu"],
    }

    # Kuyruk: öğrenci adı geçersiz → elle; öğretmen otomatik değil / telefon geçersiz → eşleştirme.
    if not ogr["gecerli"] or not kurlar:
        kuyruk = "elle"
    elif not ot["otomatik"] or not tel["gecerli"]:
        kuyruk = "eslestirme"
    else:
        kuyruk = "temiz"

    return {
        "satir_no": satir_no,
        "ham": ham,
        "norm": norm,
        "ogretmen_oneri": ot,
        "secili_ogretmen_id": ot["en_iyi"]["id"] if ot["otomatik"] and ot["en_iyi"] else None,
        "yeni_ogretmen_ad": ham.get("ogretmen_ad", "") if not ot["en_iyi"] else "",
        "kuyruk": kuyruk,
        "atla": False,   # admin bir satırı atlamak isterse
    }


def _ozet(satirlar: list[dict]) -> dict:
    from collections import Counter
    say = Counter(s["kuyruk"] for s in satirlar)
    return {
        "toplam_satir": len(satirlar),
        "temiz": say.get("temiz", 0),
        "eslestirme": say.get("eslestirme", 0),
        "elle": say.get("elle", 0),
    }


@router.post("/toplu-kayit/yukle")
async def toplu_kayit_yukle(dosya: UploadFile = File(...), sayfa: str = Form(None),
                            current_user=Depends(_YAZMA)):
    """xlsx/csv yükler → parse + normalize → TASLAK oluşturur. DB'ye kullanıcı YAZMAZ.
    Çok sayfalı xlsx'te `sayfa` verilebilir (yoksa ilk sayfa)."""
    icerik = await dosya.read()
    if len(icerik) > MAX_DOSYA_BYTE:
        raise HTTPException(status_code=400, detail="Dosya en fazla 8 MB olabilir")
    satirlar_ham, sayfalar = _parse_dosya(icerik, dosya.filename or "", sayfa)
    if not satirlar_ham:
        raise HTTPException(status_code=400, detail="Dosya boş veya okunamadı")

    baslik_atla = _baslik_mi(satirlar_ham[0])
    veri = satirlar_ham[1:] if baslik_atla else satirlar_ham
    kolon = dict(_VARSAYILAN_KOLON)
    ogretmenler = await _ogretmen_listesi()

    # Boş olmayan satırları sakla (yeniden kolon-eşleme için ham liste de tutulur).
    ham_satirlar = [s for s in veri if any(str(c).strip() for c in s)]
    satirlar = [_satir_isle(satir, kolon, ogretmenler, i + 1) for i, satir in enumerate(ham_satirlar)]

    taslak = {
        "id": str(uuid.uuid4()),
        "durum": "taslak",
        "dosya_adi": dosya.filename or "",
        "kolon_esleme": kolon,
        "baslik_atlandi": baslik_atla,
        "sayfalar": sayfalar,
        "secili_sayfa": (sayfa if (sayfa and sayfa in sayfalar) else (sayfalar[0] if sayfalar else None)),
        "ham_satirlar": ham_satirlar,
        "varsayilan_ucret": None,
        "sinif_ucret": {},       # {"3": 2500, ...} opsiyonel
        "satirlar": satirlar,
        "ozet": _ozet(satirlar),
        "yukleyen_id": current_user.get("id"),
        "yukleyen_ad": f"{current_user.get('ad','')} {current_user.get('soyad','')}".strip(),
        "tarih": datetime.utcnow().isoformat(),
    }
    await db.toplu_kayit_taslaklari.insert_one(dict(taslak))
    taslak.pop("_id", None)
    return {"taslak_id": taslak["id"], "ozet": taslak["ozet"], "kolon_esleme": kolon,
            "baslik_atlandi": baslik_atla, "satir_sayisi": len(satirlar),
            "sayfalar": sayfalar, "secili_sayfa": taslak["secili_sayfa"],
            "ogretmenler": [{"id": t["id"], "ad": f"{t.get('ad','')} {t.get('soyad','')}".strip()} for t in ogretmenler]}


@router.get("/toplu-kayit/taslak/{taslak_id}")
async def toplu_kayit_taslak(taslak_id: str, current_user=Depends(_YAZMA)):
    """Eşleştirme ekranı için tam taslak (satırlar + öneriler + özet + öğretmen listesi)."""
    t = await db.toplu_kayit_taslaklari.find_one({"id": taslak_id}, {"_id": 0, "ham_satirlar": 0})
    if not t:
        raise HTTPException(status_code=404, detail="Taslak bulunamadı")
    ogretmenler = await _ogretmen_listesi()
    t["ogretmenler"] = [{"id": o["id"], "ad": f"{o.get('ad','')} {o.get('soyad','')}".strip()} for o in ogretmenler]
    return t


@router.put("/toplu-kayit/taslak/{taslak_id}")
async def toplu_kayit_taslak_guncelle(taslak_id: str, data: dict, current_user=Depends(_YAZMA)):
    """Admin düzeltmeleri: kolon eşleme, varsayılan/sınıf ücreti, satır bazlı override
    (secili_ogretmen_id, yeni_ogretmen_ad, düzeltilmiş ad/soyad/sınıf/telefon, kuyruk, atla).
    Kolon eşleme değişirse satırlar yeniden normalize edilir."""
    t = await db.toplu_kayit_taslaklari.find_one({"id": taslak_id})
    if not t:
        raise HTTPException(status_code=404, detail="Taslak bulunamadı")
    if t.get("durum") != "taslak":
        raise HTTPException(status_code=400, detail="Yalnız 'taslak' durumundaki kayıt düzenlenebilir")

    guncelle = {}
    if "varsayilan_ucret" in data:
        try:
            guncelle["varsayilan_ucret"] = float(data["varsayilan_ucret"]) if data["varsayilan_ucret"] is not None else None
        except (TypeError, ValueError):
            raise HTTPException(status_code=422, detail="varsayilan_ucret sayısal olmalı")
    if isinstance(data.get("sinif_ucret"), dict):
        guncelle["sinif_ucret"] = {str(k): float(v) for k, v in data["sinif_ucret"].items()}

    # Kolon eşleme değişti → ham satırlardan yeniden normalize et (gerçek yeniden-eşleme).
    if isinstance(data.get("kolon_esleme"), dict):
        kolon = {**t.get("kolon_esleme", {}), **{k: int(v) for k, v in data["kolon_esleme"].items() if v is not None}}
        ogretmenler = await _ogretmen_listesi()
        ham_satirlar = t.get("ham_satirlar", [])
        guncelle["satirlar"] = [_satir_isle(satir, kolon, ogretmenler, i + 1) for i, satir in enumerate(ham_satirlar)]
        guncelle["ozet"] = _ozet(guncelle["satirlar"])
        guncelle["kolon_esleme"] = kolon

    # Satır bazlı override'lar: [{satir_no, secili_ogretmen_id?, yeni_ogretmen_ad?, kuyruk?, atla?, norm:{...}}]
    if isinstance(data.get("satir_guncelle"), list):
        harita = {s["satir_no"]: s for s in t["satirlar"]}
        for upd in data["satir_guncelle"]:
            sn = upd.get("satir_no")
            if sn not in harita:
                continue
            s = harita[sn]
            for alan in ("secili_ogretmen_id", "yeni_ogretmen_ad", "kuyruk", "atla"):
                if alan in upd:
                    s[alan] = upd[alan]
            if isinstance(upd.get("norm"), dict):
                s["norm"].update({k: v for k, v in upd["norm"].items()})
                # Öğrenci adı elle tamamlandıysa geçerli say.
                if s["norm"].get("ogrenci_ad") and s["norm"].get("ogrenci_soyad"):
                    s["norm"]["ogrenci_gecerli"] = True
        guncelle["satirlar"] = list(harita.values())
        guncelle["ozet"] = _ozet(guncelle["satirlar"])

    if not guncelle:
        raise HTTPException(status_code=400, detail="Güncellenecek alan yok")
    await db.toplu_kayit_taslaklari.update_one({"id": taslak_id}, {"$set": guncelle})
    return {"ok": True, "ozet": guncelle.get("ozet", t.get("ozet"))}


# ─────────────────────────── Uygulama (idempotent) ───────────────────────────
def _slug(ad: str, soyad: str) -> str:
    s = KN._ascii_fold(f"{ad} {soyad}").replace(" ", ".")
    s = re.sub(r"[^a-z0-9.]", "", s).strip(".")
    return s or "kullanici"


async def _benzersiz_email(taban: str, kullanilan: set) -> str:
    for i in range(0, 1000):
        email = f"{taban}@toplu.oba" if i == 0 else f"{taban}{i}@toplu.oba"
        if email not in kullanilan and not await db.users.find_one({"email": email}, {"_id": 1}):
            kullanilan.add(email)
            return email
    email = f"{taban}.{uuid.uuid4().hex[:6]}@toplu.oba"
    kullanilan.add(email)
    return email


def _kur_adi(n: int) -> str:
    return f"Kur {n}"


@router.post("/toplu-kayit/uygula/{taslak_id}")
async def toplu_kayit_uygula(taslak_id: str, dry_run: bool = Query(False), current_user=Depends(_YAZMA)):
    """Taslağı uygular: öğretmen/öğrenci/veli kullanıcıları + kur-alacak (idempotent).
    dry_run=True → hiçbir şey yazmaz, yalnız raporu döner. Aynı dosya ikinci kez
    uygulanırsa (telefon+ad+kur anahtarıyla) mükerrer OLUŞTURMAZ."""
    t = await db.toplu_kayit_taslaklari.find_one({"id": taslak_id})
    if not t:
        raise HTTPException(status_code=404, detail="Taslak bulunamadı")
    if t.get("durum") == "uygulandi" and not dry_run:
        raise HTTPException(status_code=400, detail="Bu taslak zaten uygulanmış")

    varsayilan = t.get("varsayilan_ucret")
    sinif_ucret = t.get("sinif_ucret", {}) or {}
    now = datetime.utcnow().isoformat()

    rapor = {"okundu": 0, "ogretmen_olusturuldu": 0, "ogretmen_eslesti": 0,
             "ogrenci_olusturuldu": 0, "ogrenci_eslesti": 0, "veli_olusturuldu": 0,
             "veli_eslesti": 0, "kur_alacak": 0, "elle_kuyrugu": 0, "atlanan": []}
    olusturulan = []          # {ad, soyad, rol, email, telefon, gecici_sifre}
    kullanilan_email = set()
    ogretmen_cache = {}       # ascii-ad → teacher_id
    ogretmen_gorulen = set()  # sayım için (benzersiz teacher_id)
    ogrenci_cache = {}        # (tel, ad, soyad) → student_id
    veli_cache = {}           # telefon → parent_id
    ogretmenler_liste = await _ogretmen_listesi()

    async def _ogretmen_coz(satir):
        """Dönüş: teacher_id | None."""
        tid = satir.get("secili_ogretmen_id")
        if tid:
            if tid not in ogretmen_gorulen:
                ogretmen_gorulen.add(tid); rapor["ogretmen_eslesti"] += 1
            return tid
        ham = satir.get("yeni_ogretmen_ad") or ""
        if not ham.strip():
            return None
        anah = KN._ascii_fold(re.sub(r"\b(hoca(m)?|öğretmen(im)?)\b", "", KN.tr_kucuk(ham)).strip())
        if anah in ogretmen_cache:
            return ogretmen_cache[anah]
        # Re-run idempotency: mevcut öğretmenlerle güçlü eşleşme?
        es = KN.ogretmen_eslestir(ham, ogretmenler_liste)
        if es["en_iyi"] and es["en_iyi"]["skor"] >= 0.9:
            tid = es["en_iyi"]["id"]; ogretmen_cache[anah] = tid
            if tid not in ogretmen_gorulen:
                ogretmen_gorulen.add(tid); rapor["ogretmen_eslesti"] += 1
            return tid
        # Yeni öğretmen
        tam = KN.normalize_ad(re.sub(r"\b(hoca(m)?|öğretmen(im)?)\b", "", KN.tr_kucuk(ham)).strip())
        parca = tam.split(" ", 1); ad = parca[0] or "Öğretmen"; soyad = parca[1] if len(parca) > 1 else ""
        if dry_run:
            tid = "dry-" + uuid.uuid4().hex[:12]
        else:
            email = await _benzersiz_email(_slug(ad, soyad), kullanilan_email)
            sifre = gecici_sifre_uret()
            uid = str(uuid.uuid4())
            udoc = {"id": uid, "ad": ad, "soyad": soyad, "email": email, "telefon": "",
                    "role": "teacher", "password_hash": hash_password(sifre),
                    "sifre_degistirme_zorunlu": True, "olusturma_tarihi": now, "kaynak": "toplu_kayit"}
            await db.users.insert_one(udoc)
            tid = await ogretmen_kaydi_olustur(udoc) or uid
            ogretmenler_liste.append({"id": tid, "ad": ad, "soyad": soyad})
            olusturulan.append({"ad": ad, "soyad": soyad, "rol": "teacher", "email": email, "telefon": "", "gecici_sifre": sifre})
        ogretmen_cache[anah] = tid
        rapor["ogretmen_olusturuldu"] += 1
        return tid

    for s in t["satirlar"]:
        rapor["okundu"] += 1
        n = s["norm"]
        if s.get("atla"):
            rapor["atlanan"].append({"satir_no": s["satir_no"], "sebep": "admin atladı"}); continue
        if not n.get("ogrenci_gecerli") or not n.get("kurlar"):
            rapor["elle_kuyrugu"] += 1
            rapor["atlanan"].append({"satir_no": s["satir_no"], "ogrenci": f"{n.get('ogrenci_ad','')} {n.get('ogrenci_soyad','')}".strip(),
                                     "sebep": n.get("ogrenci_sebep") or "kur/ad eksik"}); continue

        teacher_id = await _ogretmen_coz(s)

        # --- Öğrenci (idempotent: veli_tel + ad + soyad) ---
        ogr_anah = (n.get("veli_telefon") or "", n["ogrenci_ad"], n["ogrenci_soyad"])
        student_id = ogrenci_cache.get(ogr_anah)
        if not student_id:
            mevcut = await db.students.find_one(
                {"veli_telefon": n.get("veli_telefon") or "", "ad": n["ogrenci_ad"], "soyad": n["ogrenci_soyad"]}, {"id": 1})
            if mevcut:
                student_id = mevcut["id"]; rapor["ogrenci_eslesti"] += 1
            else:
                student_id = "dry-" + uuid.uuid4().hex[:12] if dry_run else str(uuid.uuid4())
                rapor["ogrenci_olusturuldu"] += 1
                if not dry_run:
                    sdoc = {"id": student_id, "ad": n["ogrenci_ad"], "soyad": n["ogrenci_soyad"],
                            "sinif": str(n.get("sinif") or ""), "veli_ad": n.get("veli_ad", ""),
                            "veli_soyad": n.get("veli_soyad", ""), "veli_telefon": n.get("veli_telefon") or "",
                            "aldigi_egitim": "", "kur": str(n["kurlar"][0]) if n["kurlar"] else "",
                            "yapilmasi_gereken_odeme": 0.0, "yapilan_odeme": 0.0, "ogretmene_yapilacak_odeme": 0.0,
                            "ogretmen_id": teacher_id, "arsivli": False, "olusturma_tarihi": now, "kaynak": "toplu_kayit"}
                    if n.get("egitim_notu"):
                        sdoc["egitim_notu"] = n["egitim_notu"]   # yalnız eğitimci rollerinin göreceği alan
                    await db.students.insert_one(sdoc)
                    if teacher_id and not str(teacher_id).startswith("dry-"):
                        await db.teachers.update_one({"id": teacher_id},
                            {"$inc": {"ogrenci_sayisi": 1}, "$addToSet": {"atanan_ogrenciler": student_id}})
                    email = await _benzersiz_email(_slug(n["ogrenci_ad"], n["ogrenci_soyad"]), kullanilan_email)
                    sifre = gecici_sifre_uret()
                    await db.users.insert_one({"id": str(uuid.uuid4()), "ad": n["ogrenci_ad"], "soyad": n["ogrenci_soyad"],
                        "email": email, "telefon": "", "role": "student", "linked_id": student_id,
                        "password_hash": hash_password(sifre), "sifre_degistirme_zorunlu": True,
                        "olusturma_tarihi": now, "kaynak": "toplu_kayit"})
                    olusturulan.append({"ad": n["ogrenci_ad"], "soyad": n["ogrenci_soyad"], "rol": "student", "email": email, "telefon": "", "gecici_sifre": sifre})
            ogrenci_cache[ogr_anah] = student_id

        # Eğitim/sağlık notu (yalnız eğitimci rollerinin gördüğü alan) — öğrenci yeni
        # ya da önceki satırdan mevcut olsun, bu satırda not varsa uygula.
        if n.get("egitim_notu") and not dry_run and not str(student_id).startswith("dry-"):
            await db.students.update_one({"id": student_id}, {"$set": {"egitim_notu": n["egitim_notu"]}})

        # --- Veli (parent user, idempotent: telefon) ---
        tel = n.get("veli_telefon")
        if tel and tel not in veli_cache:
            mevcut_p = await db.users.find_one({"role": "parent", "telefon": tel}, {"id": 1})
            if mevcut_p:
                veli_cache[tel] = mevcut_p["id"]; rapor["veli_eslesti"] += 1
            else:
                rapor["veli_olusturuldu"] += 1
                if dry_run:
                    veli_cache[tel] = "dry-" + uuid.uuid4().hex[:12]
                else:
                    email = await _benzersiz_email(_slug(n.get("veli_ad", "veli"), n.get("veli_soyad", "")), kullanilan_email)
                    sifre = gecici_sifre_uret()
                    pid = str(uuid.uuid4())
                    await db.users.insert_one({"id": pid, "ad": n.get("veli_ad") or "Veli", "soyad": n.get("veli_soyad", ""),
                        "email": email, "telefon": tel, "role": "parent", "linked_id": student_id,
                        "password_hash": hash_password(sifre), "sifre_degistirme_zorunlu": True,
                        "olusturma_tarihi": now, "kaynak": "toplu_kayit"})
                    olusturulan.append({"ad": n.get("veli_ad") or "Veli", "soyad": n.get("veli_soyad", ""), "rol": "parent", "email": email, "telefon": tel, "gecici_sifre": sifre})
                    veli_cache[tel] = pid

        # --- Kur alacakları (idempotent: ogrenci_id + kur_adi) ---
        gercek_ogrenci = not str(student_id).startswith("dry-")
        notlar = "; ".join([x for x in [n.get("aciklama"), n.get("taksit_notu")] if x])
        for kn in n["kurlar"]:
            kur_ad = _kur_adi(kn)
            if gercek_ogrenci and await db.kur_ucretleri.find_one({"ogrenci_id": student_id, "kur_adi": kur_ad}, {"_id": 1}):
                continue  # zaten var → idempotent
            ucret = float(sinif_ucret.get(str(n.get("sinif")), varsayilan or 0) or 0)
            iptal = (n.get("odeme_durumu") == "iptal")
            tutar = 0.0 if iptal else ucret
            rapor["kur_alacak"] += 1
            if dry_run or not gercek_ogrenci:
                continue
            await db.kur_ucretleri.insert_one({"id": str(uuid.uuid4()), "ogrenci_id": student_id, "kur_adi": kur_ad,
                "tutar": tutar, "baslangic_tarihi": n.get("kayit_tarihi"), "tarih": now,
                "ekleyen_id": current_user.get("id"), "kaynak": "toplu_kayit"})
            if tutar > 0:
                inc = {"yapilmasi_gereken_odeme": tutar}
                if n.get("odeme_durumu") in ("odendi", "tamamlandi"):
                    inc["yapilan_odeme"] = tutar
                await db.students.update_one({"id": student_id}, {"$inc": inc})
            if notlar:
                await db.students.update_one({"id": student_id}, {"$set": {"muhasebe_notu": notlar}})

    if not dry_run:
        await db.toplu_kayit_taslaklari.update_one({"id": taslak_id}, {"$set": {
            "durum": "uygulandi", "rapor": rapor, "olusturulan_kullanicilar": olusturulan,
            "uygulama_tarihi": now}})
    return {"dry_run": dry_run, "rapor": rapor, "olusturulan_kullanici_sayisi": len(olusturulan)}


# ─────────────────────────── Raporlar (xlsx) ───────────────────────────
def _xlsx_yanit(wb, dosya_adi: str):
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{dosya_adi}"'})


@router.get("/toplu-kayit/rapor/{taslak_id}/sifreler.xlsx")
async def rapor_sifreler(taslak_id: str, current_user=Depends(_YAZMA)):
    """Oluşturulan kullanıcı adı + geçici şifre listesi (veli/öğrenciye dağıtım için)."""
    t = await db.toplu_kayit_taslaklari.find_one({"id": taslak_id})
    if not t:
        raise HTTPException(status_code=404, detail="Taslak bulunamadı")
    import openpyxl
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Sifreler"
    ws.append(["Ad", "Soyad", "Rol", "Kullanici Adi (E-posta)", "Telefon", "Gecici Sifre"])
    for u in t.get("olusturulan_kullanicilar", []):
        ws.append([u.get("ad", ""), u.get("soyad", ""), u.get("rol", ""), u.get("email", ""), u.get("telefon", ""), u.get("gecici_sifre", "")])
    return _xlsx_yanit(wb, f"sifreler_{taslak_id[:8]}.xlsx")


@router.get("/toplu-kayit/rapor/{taslak_id}/hatalar.xlsx")
async def rapor_hatalar(taslak_id: str, current_user=Depends(_YAZMA)):
    """Atlanan / elle tamamlanacak satırların listesi (sebeplerle)."""
    t = await db.toplu_kayit_taslaklari.find_one({"id": taslak_id})
    if not t:
        raise HTTPException(status_code=404, detail="Taslak bulunamadı")
    import openpyxl
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Atlanan"
    ws.append(["Satir No", "Ogrenci", "Sebep"])
    for a in (t.get("rapor", {}) or {}).get("atlanan", []):
        ws.append([a.get("satir_no", ""), a.get("ogrenci", ""), a.get("sebep", "")])
    return _xlsx_yanit(wb, f"atlanan_{taslak_id[:8]}.xlsx")
